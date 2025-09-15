###################################################################
# 根據 doc/table.xlsx 的 kk sheet
# 擷取每個音標的 mp3 片段，合併並存到 ../audio/kk.mp3
# 把 doc/table.xlsx 的 kk sheet 的 mp3 / start / end 欄位
# 用 kk.mp3 及相對應的 start / end 取代
# 並存於 ./output/kk.xlsx 及 ../content/kk.txt 
###################################################################
# merge.py
from __future__ import annotations
import sys
from pathlib import Path
from typing import Optional, List, Tuple
from openpyxl import load_workbook
from pydub import AudioSegment
from pydub.utils import which

# ========= 路徑（相對本檔）=========
BASE = Path(__file__).resolve().parent
EXCEL_PATH   = (BASE / "../doc/kk.xlsx").resolve()
SHEET_NAME   = "kk"
HEADER_ROW   = 5  # 第5列為欄名（1-based）

AUDIO_SRC    = (BASE / "../tmp/audio").resolve()
MP3_OUT_DIR  = (BASE / "../audio").resolve()
XLSX_OUT_DIR = (BASE / "output").resolve()
TXT_OUT_DIR  = (BASE / "../content").resolve()

MP3_OUT  = MP3_OUT_DIR / "kk.mp3"
XLSX_OUT = XLSX_OUT_DIR / "kk.xlsx"
TXT_OUT  = TXT_OUT_DIR / "kk.txt"

GAP_MS = 1000  # 片段之間插入 1 秒空白

# ========= 檢查與建立輸出目錄 =========
if not EXCEL_PATH.exists():
    sys.exit(f"找不到 Excel：{EXCEL_PATH}")
for d in (MP3_OUT_DIR, XLSX_OUT_DIR, TXT_OUT_DIR):
    d.mkdir(parents=True, exist_ok=True)

if not which("ffmpeg"):
    print("⚠️ 找不到 ffmpeg（pydub 需要）。請先安裝並加入 PATH。")
if not AUDIO_SRC.exists():
    print(f"⚠️ 找不到來源音檔資料夾：{AUDIO_SRC}（仍會嘗試相對/絕對路徑）")

# ========= 讀整張 kk（保留 1~4 列與所有欄位）=========
wb = load_workbook(EXCEL_PATH, data_only=True)
if SHEET_NAME not in wb.sheetnames:
    sys.exit(f"找不到工作表：{SHEET_NAME}")
ws = wb[SHEET_NAME]
max_row, max_col = ws.max_row, ws.max_column
if HEADER_ROW > max_row:
    sys.exit(f"表格列數不足，沒有第 {HEADER_ROW} 列可當標頭")

# ========= 欄位索引（1-based）=========
def find_col(header_name: str, ci: bool = False) -> Optional[int]:
    target = header_name.strip()
    t_l = target.lower()
    for c in range(1, max_col + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str):
            vv = v.strip()
            if (ci and vv.lower() == t_l) or (not ci and vv == target):
                return c
    return None

COL_PHON = find_col("音標",  ci=False)
COL_MP3  = find_col("mp3",   ci=True)
COL_S    = find_col("start", ci=True)
COL_E    = find_col("end",   ci=True)
missing = [name for name, col in [("音標", COL_PHON), ("mp3", COL_MP3), ("start", COL_S), ("end", COL_E)] if col is None]
if missing:
    sys.exit(f"欄位缺少：{missing}（請確認第 {HEADER_ROW} 列有 音標 / mp3 / start / end）")

# ========= 解析 mp3 路徑 =========
def resolve_audio_path(name: str) -> Optional[Path]:
    if not name:
        return None
    name = str(name).strip()
    # 1) ../tmp/audio/<name>[.mp3]
    cands = [name] if name.lower().endswith(".mp3") else [name + ".mp3"]
    for nm in cands:
        p = (AUDIO_SRC / nm).resolve()
        if p.exists():
            return p
    # 2) 視為相對/絕對路徑
    for nm in [name] + cands:
        p = Path(nm).expanduser().resolve()
        if p.exists():
            return p
    # 3) 遞迴搜尋來源資料夾（完整檔名或 stem）
    if AUDIO_SRC.exists():
        stem = Path(cands[0]).stem.lower()
        for path in AUDIO_SRC.rglob("*.mp3"):
            if path.name.lower() == cands[0].lower() or path.stem.lower() == stem:
                return path.resolve()
    return None

# ========= 預掃描有效片段（先決定合併順序，好加入 1 秒 gap）=========
def to_float(x) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

valid_rows: List[Tuple[int, Path, float, float]] = []  # (row_index, path, start_s, end_s)
for r in range(HEADER_ROW + 1, max_row + 1):
    v_mp3 = ws.cell(row=r, column=COL_MP3).value
    v_s   = ws.cell(row=r, column=COL_S).value
    v_e   = ws.cell(row=r, column=COL_E).value
    mp3name = "" if v_mp3 is None else str(v_mp3).strip()
    s = to_float(v_s)
    e = to_float(v_e)
    if not mp3name or s is None or e is None or e <= s:
        continue
    apath = resolve_audio_path(mp3name)
    if apath is None:
        print(f"  ✗ 預掃描：列 {r} 找不到音檔 {mp3name}")
        continue
    valid_rows.append((r, apath, s, e))

if not valid_rows:
    print("⚠️ 沒有可處理的片段；將直接輸出 kk.xlsx 與 kk.txt（不修改內容，不輸出 kk.mp3）。")
    wb.save(XLSX_OUT)
    with open(TXT_OUT, "w", encoding="utf-8", newline="\n") as f:
        for r in range(1, ws.max_row + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                s = "" if v is None else str(v)
                s = s.replace("\t", " ").replace("\r\n", " ").replace("\n", " ")
                vals.append(s)
            f.write("\t".join(vals) + "\n")
    print(f"✅ 已輸出 Excel：{XLSX_OUT}")
    print(f"✅ 已輸出 TXT：{TXT_OUT}（TSV；與 Excel 同欄位）")
    sys.exit(0)

# ========= 合併（片段間插入 1 秒 gap）並寫回時間軸 =========
merged: AudioSegment | None = None
cur_ms = 0

print("▶ 開始合併（片段間 1 秒 gap）…")
for idx, (r, apath, s, e) in enumerate(valid_rows):
    seg = AudioSegment.from_file(apath)  # 需 ffmpeg
    s_ms = max(0, min(int(s * 1000), len(seg)))
    e_ms = max(0, min(int(e * 1000), len(seg)))
    if e_ms <= s_ms:
        print(f"  ✗ 列 {r}: 時間區間異常 start={s}, end={e}")
        continue

    clip = seg[s_ms:e_ms]

    # 在片段之前加入 1 秒 gap
    merged = AudioSegment.silent(duration=GAP_MS) if merged is None else (merged + AudioSegment.silent(duration=GAP_MS))
    cur_ms += GAP_MS

    # 合併後 kk.mp3 的時間（秒，三位小數）
    merged_s = round(cur_ms / 1000.0, 3)
    merged_e = round((cur_ms + len(clip)) / 1000.0, 3)

    # 串接
    merged = clip if merged is None else (merged + clip)
    cur_ms += len(clip)

    # # 在片段之間加入 1 秒 gap（最後一段不加）
    # if idx < len(valid_rows) - 1:
    #     merged += AudioSegment.silent(duration=GAP_MS)
    #     cur_ms += GAP_MS

    # ★ 更新工作表：把原來的 mp3/start/end 改為合併後 kk.mp3 的時間軸（已包含先前片段與 gap）
    ws.cell(row=r, column=COL_MP3, value="kk.mp3")
    ws.cell(row=r, column=COL_S,   value=(merged_s-0.25))  # start 時間點往前移 0.25 秒
    ws.cell(row=r, column=COL_E,   value=merged_e)

    print(f"  ✓ 列 {r}: {apath.name} → [{merged_s}, {merged_e}]（已考慮前面片段與 1s gap）")

# ========= 匯出 kk.mp3 =========
if merged is not None:
    merged.export(MP3_OUT, format="mp3", bitrate="192k")
    print(f"✅ 已輸出 MP3：{MP3_OUT}")
else:
    print("❌ 未產生 kk.mp3（沒有成功合併的片段）")

# ========= 匯出 kk.xlsx（完整保留 1~4 列與所有欄位）=========
wb.save(XLSX_OUT)
print(f"✅ 已輸出 Excel：{XLSX_OUT}（已更新 mp3/start/end；保留所有欄位與第 1~4 列）")

# ========= 匯出 kk.txt（TSV；內容與 kk.xlsx 同欄位）=========
def cell_to_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.replace("\t", " ").replace("\r\n", " ").replace("\n", " ")

with open(TXT_OUT, "w", encoding="utf-8", newline="\n") as f:
    for r in range(1, ws.max_row + 1):
        row_vals = [cell_to_text(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        f.write("\t".join(row_vals) + "\n")

print(f"✅ 已輸出 TXT：{TXT_OUT}（與 kk.xlsx 同欄位，TSV 格式）")






